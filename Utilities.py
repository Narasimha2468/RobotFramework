import re
from datetime import datetime
import importlib
import os
import subprocess
import json
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import psutil
import getpass
import time
import glob
from pymongo import MongoClient
import win32com.client as win32
import zipfile
import platform
import math
from cryptography.fernet import Fernet
import pandas as pd
from faker import Faker


KEY_FILE = r"D:\IBorgAgentSetup2.1.6\iborgagent\Scripts\keys\public_key.key"


def group_failed_testcases(data):
    fail_info = []

    for sublist in data:
        if sublist['Status'] == "FAILED":
            info_dict = {'test_case_id': sublist['TestCaseID'], 'test_case_desc': sublist['TestCaseDesc'], 'reason': sublist['Reason'][:-1]}
            fail_info.append(info_dict)
    return fail_info


def generate_fake_data(data_types, num_records=3):
    fake = Faker()
    fake_data = []

    for _ in range(num_records):
        record = {}
        for data_type in data_types:
            try:
                record[data_type] = getattr(fake, data_type)()
            except AttributeError:
                raise ValueError(f"Data type '{data_type}' doesn't exist. Please check https://faker.readthedocs.io/en/master/index.html to see available data types.")
        fake_data.append(record)

    return fake_data


def df_to_excel(input_directory, output_file):
    input_file_paths = []
    dfs_list = []

    try:
        if os.path.exists(output_file):
            os.remove(output_file)
    except Exception as e:
        raise FileExistsError(f"{output_file} already existes. Error overwriting it: {e}")

    for filename in os.listdir(input_directory):
        full_path = os.path.join(input_directory, filename)
        if filename.endswith(".xlsx") and os.path.isfile(full_path):
            input_file_paths.append(full_path)

    if not input_file_paths:
        raise ValueError("No .xlsx files found in the input directory.")

    for input_file_path in input_file_paths:
        try:
            df = pd.read_excel(input_file_path, engine='openpyxl')
            dfs_list.append(df)
        except Exception as e:
            raise Exception(f"Error reading file {input_file_path}: {e}")

    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            for idx, df in enumerate(dfs_list):
                startrow = 0 if writer.sheets.get('Sheet1') is None else writer.sheets['Sheet1'].max_row
                header_value = True if idx == 0 else False
                df.to_excel(writer, sheet_name='Sheet1', startrow=startrow, index=False, header=header_value)
    except Exception as e:
        raise Exception(f"Error writing to Excel file: {e}")


def compress_the_folder(folder_to_compress, output_folder):
    normalized_input_path = os.path.normpath(folder_to_compress)
    normalized_output_path = os.path.normpath(output_folder)

    if not os.path.exists(normalized_input_path):
        raise FileNotFoundError(f"The folder '{normalized_input_path}' does not exist.")

    if not os.path.exists(normalized_output_path):
        os.makedirs(normalized_output_path)

    # Get the name of the input folder
    folder_name = os.path.basename(normalized_input_path)

    # Create the 7z archive for the folder
    output_archive_path = os.path.join(normalized_output_path, f'{folder_name}.7z')
    subprocess.run(['7z', 'a', output_archive_path, normalized_input_path])


def load_key(KEY_FILE):
    with open(KEY_FILE, "rb") as file:
        key = file.read()
    return key


def encrypt_text_with_keyfile(text):
    key = load_key(KEY_FILE)
    cipher = Fernet(key)
    encrypted_text = cipher.encrypt(text.encode())
    return encrypted_text


def decrypt_text_with_keyfile(encrypted_text):
    key = load_key(KEY_FILE)
    cipher = Fernet(key)
    decrypted_text = cipher.decrypt(encrypted_text).decode()
    return decrypted_text


def evaluate_expression(expression):
    try:
        replacements = {
            'sine': 'math.sin',
            'sin': 'math.sin',
            'cosine': 'math.cos',
            'cos': 'math.cos',
            'tangent': 'math.tan',
            'tan': 'math.tan',
            'cotangent': '1/math.tan',
            'cot': '1/math.tan',
            'secant': '1/math.cos',
            'sec': '1/math.cos',
            'cosecant': '1/math.sin',
            'csc': '1/math.sin',
            'log': 'math.log',
            '^': '**'
        }
        for key, value in replacements.items():
            expression = expression.replace(key, value)

        result = eval(expression)
        return result
    except Exception as e:
        return f"Error: {e}"


def get_disk_space():
    # Disk space calculation using subprocess
    result = subprocess.run(['wmic', 'logicaldisk', 'get', 'size,freespace,caption'], capture_output=True, text=True)
    lines = result.stdout.strip().split('\n')
    disk_space = {}
    for line in lines[1:]:
        parts = line.split()
        if len(parts) >= 3:
            drive = parts[0]
            total_space = int(parts[1]) // (1024 ** 3)  # Convert bytes to GB
            free_space = int(parts[2]) // (1024 ** 3)   # Convert bytes to GB
            disk_space[drive] = {
                'Total Disk Space': total_space,
                'Free Disk Space': free_space
            }
    return disk_space

def get_system_info():
    system_info = {
        'Operating System': platform.system(),
        'Python Version': platform.python_version(),
        'Hardware Architecture': platform.machine(),
        'CPU Usage': f"{psutil.cpu_percent(interval=1)}%",
        'Physical Cores': psutil.cpu_count(logical=False),
        'Total Cores': psutil.cpu_count(logical=True),
        'Total Memory': f"{psutil.virtual_memory().total / (1024 ** 3):.2f} GB",
        'Available Memory': f"{psutil.virtual_memory().available / (1024 ** 3):.2f} GB",
        'Used Memory': f"{psutil.virtual_memory().used / (1024 ** 3):.2f} GB",
        'Memory Usage Percentage': f"{psutil.virtual_memory().percent}%",
        'Disk Space': get_disk_space()
    }
    return system_info


def unit_converter(total_size):
    units = ['bytes', 'KB', 'MB', 'GB']
    divisor = 1024
    for unit in units:
        if total_size < divisor or unit == units[-1]:
            return f"{total_size:.2f} {unit}"
        total_size /= divisor


def get_size(start_path):
    total_size = 0
    try:
        if os.path.isfile(start_path):
            return os.path.getsize(start_path)
        elif os.path.isdir(start_path):
            for dirpath, _, filenames in os.walk(start_path):
                for filename in filenames:
                    filepath = os.path.join(dirpath, filename)
                    try:
                        file_size = os.path.getsize(filepath)
                        total_size += file_size
                    except OSError as e:
                        raise OSError(f"Error accessing file '{filepath}': {e}")
        return unit_converter(total_size)
    except OSError as e:
        raise OSError(f"Error accessing path '{start_path}': {e}")


def convert_docx_to_pdf(input_docx_path, output_pdf_path):
    wdFormatPDF = 17
    input_docx_path = os.path.abspath(input_docx_path)
    output_pdf_path = os.path.abspath(output_pdf_path)
    word = win32.Dispatch('Word.Application')
    try:
        doc = word.Documents.Open(input_docx_path)
        doc.SaveAs(output_pdf_path, FileFormat=wdFormatPDF)
        doc.Close()
        word.Quit()
        return True
    except Exception as e:
        if 'word' in locals():
            word.Quit()
        raise Exception(f"Error converting DOCX to PDF: {e}")


def convert_pdf_to_docx(pdf_file, docx_file, password=None):
    pdf_file = os.path.abspath(pdf_file)
    docx_file = os.path.abspath(docx_file)
    word = win32.Dispatch("Word.Application")
    word.Visible = False

    try:
        wb = word.Documents.Open(pdf_file, Password=password)
        wb.SaveAs(docx_file, FileFormat=16)  # FileFormat 16 is for DOCX format
        wb.Close()
        word.Quit()
        return True
    except Exception as e:
        if 'word' in locals():
            word.Quit()
        raise Exception(f"Error converting PDF to DOCX: {e}")


def convert_xls_to_xlsx(file_path, output_path):
   output_path = output_path.replace(".xls", ".xlsx")
   try:
        excel = win32.gencache.EnsureDispatch("Excel.Application")
        wb = excel.Workbooks.Open(file_path)
        wb.SaveAs(output_path, FileFormat=51)
        wb.Close(True)
        excel.Application.Quit()
        return output_path
   except Exception as e:
        raise Exception(f"Error: {e}")


def unzip_the_given_file(zip_file_path, extract_dir, password=None):
    try:
        with zipfile.ZipFile(zip_file_path, 'r') as zip_ref:
            if password:
                zip_ref.setpassword(password.encode('utf-8'))
            zip_ref.extractall(extract_dir)
        print(f"Successfully extracted '{zip_file_path}' to '{extract_dir}'")
    except FileNotFoundError:
        raise FileNotFoundError(f"File '{zip_file_path}' not found.")
    except zipfile.BadZipFile:
        raise zipfile.BadZipFile(f"File '{zip_file_path}' is not a valid zip file.")
    except Exception as e:
        raise Exception(f"An error occurred while extracting '{zip_file_path}': {e}")


def get_test_data_from_db(field_name):
    client = None
    try:
        client = MongoClient('mongodb://localhost:27017/')
        database = client['Excel_Driven_Framework_DB']
        collection = database['Test_Repo']

        result = collection.find_one({}, {field_name: 1})

        if result and field_name in result:
            return result[field_name]
        else:
            return None
    except Exception as e:
        raise Exception(f"An error occurred while fetching data from the database: {e}")
    finally:
        if client:
            client.close()


def list_running_processes():
    try:
        return [proc.as_dict(attrs=['pid', 'name', 'username', 'status', 'cpu_percent', 'memory_percent']) for proc in psutil.process_iter()]
    except Exception as e:
        return f"Error: {e}"


def get_process_info(process_name):
    try:
        for proc in psutil.process_iter():
            if proc.name() == process_name:
                return proc.as_dict(attrs=['pid', 'name', 'username', 'status', 'cpu_percent', 'memory_percent'])
        return None
    except Exception as e:
        return f"Error: {e}"


def is_process_running(process_name):
    try:
        for proc in psutil.process_iter():
            if proc.name() == process_name:
                return 'Running'
        return 'Not Running'
    except Exception as e:
        return f"Error: {e}"


def kill_process_by_name(process_name):
    try:
        for process in psutil.process_iter(attrs=['pid', 'name']):
            if process_name.upper() in process.info['name'].upper():
                psutil.Process(process.info['pid']).terminate()
    except psutil.NoSuchProcess:
        raise psutil.NoSuchProcess(f"No such process with name '{process_name}' found.")
    except psutil.AccessDenied:
        raise psutil.AccessDenied(f"Access denied while terminating process with name '{process_name}'.")
    except Exception as e:
        raise Exception(f"An error occurred while terminating process with name '{process_name}': {e}")


def fetch_excel_file():
    directory = os.path.expanduser('~\\Downloads')
    pattern = os.path.join(directory, '*Test_Data*.xlsx')
    matching_files = glob.glob(pattern)
    matching_files.sort(key=os.path.getmtime, reverse=True)
    latest_file = matching_files if matching_files else []
    return latest_file


def adb_start_screen_recording(device_file_path, max_time=180):
    adb_path = os.environ.get('adb', f'C:\\Users\\{getpass.getuser()}\\AppData\\Local\\Android\\Sdk\\platform-tools\\adb.exe')
    if not os.path.exists(adb_path):
        raise FileNotFoundError(f"The path '{adb_path}' does not exist.")
    record_command = f'"{adb_path}" shell screenrecord --time-limit {max_time} "{device_file_path}"'
    subprocess.Popen(record_command, shell=True)
    print("Recording started")


def adb_stop_screen_recording(device_file_path, save_path):
    adb_path = os.environ.get('adb', f'C:\\Users\\{getpass.getuser()}\\AppData\\Local\\Android\\Sdk\\platform-tools\\adb.exe')
    if not os.path.exists(adb_path):
        raise FileNotFoundError(f"The path '{adb_path}' does not exist.")
    subprocess.run([adb_path, 'shell', 'pkill', '-l', 'SIGINT', 'screenrecord'], shell=True)
    pull_command = f'"{adb_path}" pull "{device_file_path}" "{save_path}"'
    subprocess.Popen(pull_command, shell=True)
    time.sleep(10)
    print("Recording stopped")


def check_and_install_module(module_name):
    try:
        importlib.import_module(module_name)
    except ImportError:
        try:
            subprocess.run(['pip', 'install', module_name], check=True)
        except Exception as e:
            raise e

def compress_execution_summary_with_7zip(base_path, timestamp):
    # base_path = r'.\Execution Summary'
    dynamic_folder_name = f'Execution_Summary_{timestamp}'
    folder_to_compress = os.path.join(base_path, dynamic_folder_name)
    normalized_path = os.path.normpath(folder_to_compress)
    output_folder = os.path.join(normalized_path,'Compressed Archive')

    # Ensure the output folder exists or raise an error
    if not os.path.exists(normalized_path):
        raise FileNotFoundError(f"The folder '{normalized_path}' does not exist.")

    # Create the 7z archive for the folder
    output_archive_path = os.path.join(output_folder, f'{dynamic_folder_name}.7z')
    subprocess.run(['7z', 'a', output_archive_path, normalized_path])
    return True


def transfer_files(file_path, destination_path,adb_executable=r'C:\platform-tools\adb.exe'):
    adb_delete_command = f'"{adb_executable}" shell rm -r "{destination_path}/*"'
    adb_file_transfer_command = f'"{adb_executable}" push "{file_path}" "{destination_path}"'

    try:
        subprocess.run(adb_delete_command, shell=True, check=True)
        subprocess.run(adb_file_transfer_command, shell=True, check=True)
    except:
        subprocess.run(adb_file_transfer_command, shell=True, check=True)


def convert_string_to_json(input_strings, convert_to_lower=True):
    key_value_pairs = [pair.split('=', 1) for pair in ','.join(input_strings).split(',')]
    data_dict = {key.lower().strip() if convert_to_lower else key.strip(): value.strip() for key, value in key_value_pairs}
    return data_dict


def convert_to_12hr_format(time_24hr):
    time_object = datetime.strptime(time_24hr, '%H:%M:%S')
    time_12hr = time_object.strftime('%I:%M:%S %p')
    return time_12hr

def convert_to_minutes_or_seconds(time_24hr):
    if isinstance(time_24hr, str):
        time_object = datetime.strptime(time_24hr, '%H:%M:%S')
        total_seconds = time_object.hour * 3600 + time_object.minute * 60 + time_object.second
    else:
        total_seconds = time_24hr

    if total_seconds < 60:
        return f"{int(total_seconds)} sec"
    else:
        minutes = int(total_seconds // 60)
        seconds = int(total_seconds % 60)
        if seconds == 0:
            return f"{minutes} min"
        else:
            return f"{minutes} min {seconds} sec"

def remove_keys_from_json(json_data: dict, keys_to_remove: str):
    removed_pairs = {}
    keys_list = keys_to_remove.split(',')
    for key in keys_list:
        key = key.strip()
        if key in json_data:
            removed_pairs[key] = json_data.pop(key)

    json_data_list = [f"{key}={value}" for key, value in json_data.items()]
    return json_data_list, removed_pairs


def add_to_json(section, key, value, overwrite=None, json_file_path=r".\Test Suite Info\TestSuiteInfo.json"):
    with open(json_file_path, 'r') as json_file:
        json_data = json.load(json_file)
    section = section.lower()
    if section in ["query_results", "excel_data", "file_or_dir_size", "math_results", "process_status", "process_info", "all_running_processes", "python_execution_results", "javascript_execution_results", "encrypted_data", "decrypted_data", "fake_test_data"]:
        existing_keys = [k for k in json_data['query_results'].keys() if re.match(f"{key}(\d+)", k)]
        max_number = max([int(re.match(f"{key}(\d+)", k).group(1)) for k in existing_keys]) if existing_keys else 0
        key = f"{key}{max_number + 1}"
    elif section in ["browser_count", "system_info"]:
        current_count = json_data[section]
        json_data[section]= current_count + int(value)
        updated_json = json.dumps(json_data, indent=2)
        with open(json_file_path, 'w') as json_file:
            json_file.write(updated_json)
        return
    elif section.startswith("group"):
        if isinstance(value, list):
            json_data[section][key]= value
        else:
            current_count = json_data[section][key]
            json_data[section][key]= current_count + int(value)
        updated_json = json.dumps(json_data, indent=2)
        with open(json_file_path, 'w') as json_file:
            json_file.write(updated_json)
        return
    else:
        if overwrite.lower() == 'false':
            if key in json_data[section]:
                raise ValueError(f"Variable '{key}' already exists.")

    json_data[section][key] = value
    updated_json = json.dumps(json_data, indent=2)

    with open(json_file_path, 'w') as json_file:
        json_file.write(updated_json)


def calculate_execution_time(log_file_path):
    with open(log_file_path, 'r') as log_file:
        log_lines = log_file.readlines()

    test_cases = {}

    for line in log_lines:
        match_start = re.match(r'\[(.*?)\] - \[INFO\] - Testcase \'(.*?)\' execution started', line)
        match_end = re.match(r'\[(.*?)\] - \[INFO\] - Testcase \'(.*?)\' execution ended', line)

        if match_start:
            timestamp, testcase_name = match_start.groups()
            start_time = datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S.%f')
            test_cases[testcase_name] = {'start_time': start_time}
        elif match_end:
            timestamp, testcase_name = match_end.groups()
            end_time = datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S.%f')
            elapsed_time = end_time - test_cases[testcase_name]['start_time']
            test_cases[testcase_name]['elapsed_time'] = elapsed_time

    for testcase_name, details in test_cases.items():
        if 'elapsed_time' not in details:
            details['elapsed_time'] = None

    # Format start_time and elapsed_time
    for testcase_name, details in test_cases.items():
        details['start_time'] = convert_to_12hr_format(details['start_time'].strftime('%H:%M:%S'))
        if details['elapsed_time'] is not None:
            details['elapsed_time'] = convert_to_minutes_or_seconds(str(details['elapsed_time']).split(".")[0])

    return test_cases


def format_testsuite_info(sheet):
    start_row = 1
    column_letters = ['A', 'B']
    fill_color = 'C5D9F1'

    # Create a fill object
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    for column_letter in column_letters:
        # Find the first empty cell in column A and B
        end_row = start_row
        while sheet[column_letter + str(end_row)].value is not None:
            end_row += 1

        # Apply the fill to the specified range
        for row in range(start_row, end_row):
            cell = sheet[column_letter + str(row)]
            cell.fill = fill

    # Specify the border style
    border_style = Side(style='thin')

    # Create a border object
    border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

    # Find the first row with empty cells in columns A and B
    end_row = start_row
    while sheet['A' + str(end_row)].value is not None and sheet['B' + str(end_row)].value is not None:
        end_row += 1

    # Apply the border to columns A and B until the first empty cell
    for row in range(start_row, end_row):
        for col_letter in ['A', 'B']:
            cell = sheet[col_letter + str(row)]
            cell.border = border
    return end_row


def cell_border(worksheet, start_row, excel_name, excel_file_path=None):
    global start_column, end_column
    if excel_file_path is not None:
        workbook = openpyxl.load_workbook(excel_file_path)
        worksheet = workbook.active
        start_column, end_column = ('A', 'M')
        historical_report = True
    else:
        historical_report = False
        start_column, end_column = ('A', 'H') if excel_name == "TestCaseSummary" else ('A', 'J')
    max_row = 0
    start_cell = f'A{start_row}'

    for column in range(ord(start_column), ord(end_column) + 1):
        for row in worksheet.iter_rows(min_row=start_row, min_col=column, max_col=column):
            for cell in row:
                if cell.value is not None:
                    max_row = max(max_row, cell.row)

    # Create border style
    border_style = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin')
    )

    # Apply borders to the range
    for row in worksheet[start_cell:f'{end_column}{max_row}']:
        for cell in row:
            cell.border = border_style
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Apply background color (blue) to the first row
    blue_fill = PatternFill(start_color='C4D79B', end_color='C4D79B', fill_type='solid')
    for cell in worksheet[start_column+str(start_row):end_column+str(start_row)][0]:
        cell.fill = blue_fill

    if historical_report: workbook.save(excel_file_path)


def apply_filter_to_row(sheet, row_number):
    # Specify the range for the filter to cover columns A to F in the given row
    filter_range = f"{start_column}{row_number}:{end_column}{row_number}"
    # Add filter to the specified range
    sheet.auto_filter.ref = filter_range


def set_width_and_wrap_text(sheet, target_columns):
    for column in sheet.columns:
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except TypeError:
                pass

        adjusted_width = max_length + 5
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    for target_column in target_columns:
        sheet.column_dimensions[openpyxl.utils.get_column_letter(target_column['column'])].width = target_column['width']

        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=target_column['column'], max_col=target_column['column']):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')


def remove_consecutive_duplicates(sheet, column_numbers, excel_name):
    # Iterate through each column in the list
    for column_number in column_numbers:
        rows_to_clear = []

        # Iterate through rows from top to bottom
        for row in range(10, sheet.max_row + 1):
            current_value = sheet.cell(row=row, column=column_number).value
            next_value = sheet.cell(row=row + 1, column=column_number).value if row < sheet.max_row else None

            # If the current cell value is equal to the cell below, add the row to the list
            if current_value == next_value:
                rows_to_clear.append(row + 1)

        # Clear the marked rows
        for row_to_clear in rows_to_clear:
            sheet.cell(row=row_to_clear, column=column_number, value='')

        # Assign sequential numbers in column 1 for non-empty cells in the current column
        serial_number = 1
        for row in range(10, sheet.max_row + 1):
            current_value = sheet.cell(row=row, column=column_number).value

            # If the current cell value is non-empty, assign a sequential number in column 1
            if current_value:
                sheet.cell(row=row, column=1, value=serial_number)
                serial_number += 1

        status_column_index = 6 if excel_name == "TestCaseSummary" else 8
        for row in sheet.iter_rows(min_row=10, min_col=status_column_index, max_col=status_column_index):
            for cell in row:
                if cell.value == "FAILED":
                    cell.font = Font(color="FF0000")  # Red
                elif  cell.value == "PASSED":
                    cell.font = Font(color="55AB69")  # Green
                else:
                    cell.font = Font(color="D78E2D")  # Yellow


def format_report(excel_file_path, excel_name):
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active
    column_numbers = [2, 3]
    target_columns_summary = [
    {'column': 2, 'width': 25},
    {'column': 3, 'width': 30},
    {'column': 8, 'width': 35}]
    target_columns_step = [
    {'column': 2, 'width': 25},
    {'column': 3, 'width': 30},
    {'column': 10, 'width': 35}]
    # sheet = workbook[sheet_name]

    start_row = 9
    format_testsuite_info(sheet)
    cell_border(sheet, start_row, excel_name)
    apply_filter_to_row(sheet,start_row)
    if excel_name == "TestCaseSummary":
       set_width_and_wrap_text(sheet, target_columns_summary)
    else:
       set_width_and_wrap_text(sheet, target_columns_step)
    remove_consecutive_duplicates(sheet,column_numbers,excel_name)
    workbook.save(excel_file_path)

